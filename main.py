import importlib
import openpyxl
import pandas as pd
import os
from fpdf import FPDF
from tqdm import tqdm
import seaborn as sns

Input_path ="data.xlsx"

class Main():
# read all data in file
    datafile = pd.read_excel(Input_path)

# read all data for students
    data = datafile.loc[datafile['Email'].notnull(), ['Email', 'Name', 'HW1', 'HW2', 'First', 'Second', 'final', 'total Grade']]

#  get the specific data for each student
    names = data['Name']
    Hw1 = data['HW1']
    Hw2 = data['HW2']
    First = data['First']
    Second = data['Second']
    Final = data['final']
    Total = data['total Grade']

# used for read a dynamic tasks on excel file
    hw1 = datafile.columns[3]
    hw2 = datafile.columns[4]
    First = datafile.columns[5]
    Second = datafile.columns[6]
    final = datafile.columns[7]

#  Make a new Class for Genrate a PDF
    def Check(self, value):
        if value == 0 or str(value) == '':
            value = "Miss or Did not submit"
            return value
        elif value > 0:
            return str(int(value))

#----------------------------------------------------------------------
#  Make a new Class for Genrate a PDF
class PDF(FPDF):
    def __init__(self):
        super().__init__()
    # add footer that contain Page number
    def footer(self):
        self.set_y(-15)
        self.set_font('Arial', '', 12)
        self.cell(0, 8, f'Page {self.page_no()}', 0, 0, 'C')

#---------------------------------------------------------
# used for read a fake image that i save on it and remove it after that for all charts I have and for all Students
file_img= './images/image.png'
file_img2= './images/image2.png'
file_img3= './images/image3.png'

# Define My object
main = Main()
#creat output and image files
if not os.path.exists("output") or not os.path.exists("images") :
    os.mkdir("./output")
    os.mkdir("./images")
#----------------------------------------------------------------------
for i in tqdm(range(len(main.data))):
    # check for folder that saved in !
# PDF creation for Each Student that inherit from PDF Class
    pdf = PDF() # Initializing object from my Class
    pdf.add_page() # Adding new page Page 1
    #  get the specific data for each student  in the first Class Called Main()
    name = main.names.iloc[i]
    Hw1 = main.Hw1.iloc[i]
    Hw2 = main.Hw2.iloc[i]
    First = main.First.iloc[i]
    Second = main.Second.iloc[i]
    final = main.final.iloc[i]
    Total = main.Total.iloc[i]

    # Title of Pdf Header for Each Student With his name
    pdf.set_fill_color(200, 0, 0)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Arial", size=20, style="B")
    pdf.set_xy(5, 5)
    text = 'Details of Student : ' + name  # get the name of each student
    pdf.cell(0, 12.5, txt=text, ln=1, align="C", fill=True)

    # Text box of header
    pdf.set_fill_color(0, 0, 200)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Arial", size=15, style='B')
    pdf.set_xy(5, 20)
    pdf.cell(105, 10, txt='Grades in each of the course activities : ', ln=1, align="L", fill=True)
